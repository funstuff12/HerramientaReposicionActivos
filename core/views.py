from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
import openpyxl
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.db.models import Q, Sum
from django.core.exceptions import ValidationError
from django.core.serializers.json import DjangoJSONEncoder
from django.utils.safestring import mark_safe
from django.contrib import messages
from django.utils.dateparse import parse_date
from .models import Registro, Cliente, Proveedor, Maquina, AnalisisComparativo, FlujoCaja, TablaAmortizacion
from .forms import RegistroForm, MaquinaForm
from django.core.serializers import serialize
from decimal import Decimal
from datetime import datetime, date, timedelta
import json
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
import logging
import traceback

def index(request):
    return render(request, 'index.html')

# Vista para la sección de Tesorería
def vista_tesoreria(request):
    # Aquí puedes agregar la lógica para tu página de tesorería
    return render(request, 'tesoreria.html')

def vista_maquinaria(request):
    query = request.GET.get('q', '')
    
    if query:
        maquinas = Maquina.objects.filter(
            Q(nombre__icontains=query) | 
            Q(numero_serie__icontains=query)
        )
    else:
        maquinas = Maquina.objects.all()
    
    # Cambia esta línea para usar el template correcto
    return render(request, 'maquinaria.html', {  # o el nombre que tengas
        'maquinas': maquinas,
        'query': query
    })

def maquinaria_eliminar(request, maquina_id):
    """Eliminar máquina existente"""
    print(f"=== DEBUG ELIMINAR MAQUINARIA ===")
    print(f"Método HTTP: {request.method}")  
    print(f"Máquina ID recibido: {maquina_id}")
    print(f"URL completa: {request.get_full_path()}")
    
    if request.method == 'POST':
        print("Es un POST request - procediendo con eliminación")
        try:
            maquina = get_object_or_404(Maquina, id=maquina_id)
            print(f"Máquina encontrada: {maquina.nombre} (ID: {maquina.id})")
            
            nombre_maquina = maquina.nombre
            maquina.delete()
            print(f"Máquina {nombre_maquina} eliminada exitosamente")
            
            messages.success(request, f'Máquina "{nombre_maquina}" eliminada exitosamente')
            return redirect('maquinaria')
            
        except Exception as e:
            print(f"Error al eliminar máquina: {str(e)}")
            messages.error(request, f'Error al eliminar máquina: {str(e)}')
            return redirect('maquinaria')
    else:
        print(f"Método no permitido: {request.method}")
        return HttpResponseNotAllowed(['POST'])

# Configurar logging
logger = logging.getLogger(__name__)

def vista_crear_maquinaria(request):
    print(f"🔍 REQUEST METHOD: {request.method}")
    
    if request.method == 'POST':
        print("📝 POST REQUEST - Datos recibidos:")
        
        # Debug: Imprimir todos los datos POST
        for key, value in request.POST.items():
            print(f"   {key}: {value}")
        
        form = MaquinaForm(request.POST)
        
        print(f"🔍 FORM IS_VALID: {form.is_valid()}")
        
        if form.is_valid():
            try:
                print("✅ Formulario válido, intentando guardar...")
                
                # Usar transacción para asegurar consistencia
                with transaction.atomic():
                    maquina = form.save()
                    print(f"✅ Máquina guardada exitosamente con ID: {maquina.id}")
                    
                messages.success(request, f'Máquina "{maquina.nombre}" creada exitosamente.')
                return redirect('maquinaria')
                
            except Exception as e:
                print(f"❌ ERROR AL GUARDAR: {str(e)}")
                print(f"❌ TRACEBACK: {traceback.format_exc()}")
                messages.error(request, f'Error al guardar la máquina: {str(e)}')
        else:
            print("❌ FORMULARIO NO VÁLIDO - ERRORES:")
            for field, errors in form.errors.items():
                print(f"   Campo '{field}': {errors}")
                
            # Agregar mensajes de error para mostrar al usuario
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        print("📄 GET REQUEST - Mostrando formulario vacío")
        form = MaquinaForm()

    context = {
        'form': form
    }
    return render(request, 'crear_maquinaria.html', context)

def editar_maquina(request, id):
    """
    Vista para editar una máquina existente
    """
    maquina = get_object_or_404(Maquina, id=id)
    
    if request.method == 'POST':
        form = MaquinaForm(request.POST, instance=maquina)
        if form.is_valid():
            form.save()
            messages.success(request, f'La máquina "{maquina.nombre}" ha sido actualizada exitosamente.')
            return redirect('maquinaria')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = MaquinaForm(instance=maquina)
    
    return render(request, 'editar_maquina.html', {
        'form': form,
        'maquina': maquina
    })

def comparar_maquina(request):
    """Vista principal para el análisis financiero"""
    context = {}
    
    # Si viene maquina_id por parámetro GET
    maquina_id = request.GET.get('maquina_id')
    
    if maquina_id:
        try:
            maquina = get_object_or_404(Maquina, id=maquina_id)
            context['maquina_preseleccionada'] = {
                'id': str(maquina.id),
                'nombre': maquina.nombre,
                'tipo': maquina.tipo
            }
        except:
            # Si el ID no es válido, simplemente no preseleccionar nada
            pass
    
    return render(request, 'comparar.html', context)

def api_maquinas_por_tipo(request, tipo):
    """API para obtener máquinas por tipo (Defender/Challenger)"""
    if tipo not in ['Defender', 'Challenger']:
        return JsonResponse({'error': 'Tipo inválido'}, status=400)
    
    maquinas = Maquina.objects.filter(tipo=tipo).values(
        'id', 'nombre', 'numero_serie', 'date_in_service', 'criticality_ranking'
    )
    
    return JsonResponse(list(maquinas), safe=False)

def api_maquina_detalle(request, id):
    """API para obtener detalles completos de una máquina"""
    try:
        maquina = get_object_or_404(Maquina, id=id)
        
        # Función helper para convertir Decimal y manejar nulls
        def safe_decimal_to_float(value):
            if value is None:
                return 0.0
            if isinstance(value, Decimal):  # ← Cambiar Decimal por Decimal
                return float(value)
            return float(value) if value else 0.0
        
        data = {
            'id': str(maquina.id),
            'nombre': maquina.nombre,
            'tipo': maquina.tipo,
            'numero_serie': maquina.numero_serie or '',
            'criticality_ranking': safe_decimal_to_float(maquina.criticality_ranking),
            'availability': safe_decimal_to_float(maquina.availability),
            'date_in_service': maquina.date_in_service.isoformat() if maquina.date_in_service else None,
            
            # Costos de adquisición - CORREGIDO
            'purchase_price': safe_decimal_to_float(maquina.purchase_price),
            'installation_and_training_cost': safe_decimal_to_float(maquina.installation_and_training_cost),
            'setup_costs': safe_decimal_to_float(maquina.setup_costs),
            'current_resale_value': safe_decimal_to_float(maquina.current_resale_value),
            'salvage_value': safe_decimal_to_float(maquina.salvage_value),
            'acquisition_cost': safe_decimal_to_float(maquina.acquisition_cost),  # CORREGIDO
            'book_value': safe_decimal_to_float(maquina.book_value),
            
            # Costos operativos - CORREGIDO
            'annual_maintenance_labor_parts': safe_decimal_to_float(maquina.annual_maintenance_labor_parts),
            'initial_monthly_maintenance_cost': safe_decimal_to_float(maquina.initial_monthly_maintenance_cost),
            'maintenance_cost_gradient': safe_decimal_to_float(maquina.maintenance_cost_gradient),
            'cost_of_downtime': safe_decimal_to_float(maquina.cost_of_downtime),
            'operator_labor_cost': safe_decimal_to_float(maquina.operator_labor_cost),
            'energy_consumption': safe_decimal_to_float(maquina.energy_consumption),
            'energy_cost': safe_decimal_to_float(maquina.energy_cost),
            
            # Producción y vida útil - CORREGIDO
            'useful_life': maquina.useful_life or 120,  # Default 120 meses
            'monthly_operating_hours': safe_decimal_to_float(maquina.monthly_operating_hours),
            'production_rate': safe_decimal_to_float(maquina.production_rate),
            'production_rate_units': maquina.production_rate_units or '',

        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def guardar_analisis(request):
    """Guardar un nuevo análisis completo"""
    try:
        data = json.loads(request.body)
        
        # Obtener las máquinas
        defender = Maquina.objects.get(id=data['defender_id'])
        challenger = Maquina.objects.get(id=data['challenger_id'])
        
        # Crear el análisis - Convertir todos los valores a Decimal
        analisis = AnalisisComparativo.objects.create(
            nombre_analisis=data['nombre_analisis'],
            defender=defender,
            challenger=challenger,
            wacc=Decimal(str(data['wacc'])),  # Convertir float a Decimal
            tax_rate=Decimal(str(data['tax_rate'])), 
            financing_rate=Decimal(str(data['financing_rate'])),
            financing_months=int(data['financing_months']),
            pv_defender=Decimal(str(data['pv_defender'])),
            eac_defender=Decimal(str(data['eac_defender'])),
            pv_challenger=Decimal(str(data['pv_challenger'])),
            eac_challenger=Decimal(str(data['eac_challenger'])),
            recomendacion=data['recomendacion']
        )
        
        # Guardar flujos de caja del Defender
        for flujo_data in data.get('flujos_defender', []):
            FlujoCaja.objects.create(
                analisis=analisis,
                tipo_equipo='Defender',
                año=int(flujo_data['year']),
                cash_flow_bruto=Decimal(str(flujo_data['cashFlow'])),
                depreciacion=Decimal('0'),  # Calcular si es necesario
                tax_shield=Decimal('0'),    # Calcular si es necesario
                after_tax_cash_flow=Decimal(str(flujo_data['afterTaxCashFlow'])),
                present_value=Decimal(str(flujo_data['pv']))
            )
        
        # Guardar flujos de caja del Challenger
        for flujo_data in data.get('flujos_challenger', []):
            FlujoCaja.objects.create(
                analisis=analisis,
                tipo_equipo='Challenger',
                año=int(flujo_data['year']),
                cash_flow_bruto=Decimal(str(flujo_data['cashFlow'])),
                depreciacion=Decimal('0'),  # Calcular si es necesario
                tax_shield=Decimal('0'),    # Calcular si es necesario
                after_tax_cash_flow=Decimal(str(flujo_data['afterTaxCashFlow'])),
                present_value=Decimal(str(flujo_data['pv']))
            )
        
        # Generar y guardar tabla de amortización
        generar_tabla_amortizacion(analisis)
        
        return JsonResponse({
            'success': True,
            'analisis_id': str(analisis.id),
            'message': 'Análisis guardado exitosamente'
        })
        
    except Maquina.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Máquina no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def generar_tabla_amortizacion(analisis):
    """Generar tabla de amortización para el análisis"""
    # Calcular el monto del préstamo (costo inicial del challenger)
    challenger = analisis.challenger
    P = Decimal('0')
    
    # Sumar todos los costos - asegurándonos de que sean Decimal
    if challenger.purchase_price:
        P += challenger.purchase_price
    if challenger.installation_and_training_cost:
        P += challenger.installation_and_training_cost
    if challenger.setup_costs:
        P += challenger.setup_costs
    
    # Parámetros del préstamo - convertir a Decimal
    r = analisis.financing_rate / Decimal('100') / Decimal('12')  # Tasa mensual
    n = analisis.financing_months
    
    if n <= 0 or P <= 0:
        return
    
    # Calcular pago mensual
    if r == 0:
        payment = P / Decimal(str(n))
    else:
        # Fórmula de pago mensual con Decimal
        factor = (Decimal('1') + r) ** n
        payment = P * (r * factor) / (factor - Decimal('1'))
    
    # Generar tabla mes a mes
    balance = P
    
    for mes in range(1, n + 1):
        interest_payment = balance * r
        principal_payment = payment - interest_payment
        final_balance = balance - principal_payment
        
        TablaAmortizacion.objects.create(
            analisis=analisis,
            mes=mes,
            balance_inicial=balance,
            pago_mensual=payment,
            pago_principal=principal_payment,
            pago_interes=interest_payment,
            balance_final=max(Decimal('0'), final_balance)  # No permitir balance negativo
        )
        
        balance = final_balance
        if balance <= 0:
            break

def calcular_analisis_completo(analisis):
    """Función para calcular el análisis financiero completo"""
    defender = analisis.defender
    challenger = analisis.challenger
    wacc = float(analisis.wacc)
    tax_rate = float(analisis.tax_rate)
    
    # Análisis Defender
    book_value = float(defender.acquisition_cost or defender.purchase_price or 0)
    resale_value = float(defender.current_resale_value or 0)
    tax_on_sale = (resale_value - book_value) * tax_rate
    after_tax_opportunity_cost = resale_value - tax_on_sale
    
    annual_op_ex_defender = float(defender.annual_maintenance_labor_parts or 0) + \
                          (float(defender.operator_labor_cost or 0) * float(defender.monthly_operating_hours or 0) * 12)
    
    life_years_defender = (defender.useful_life or 120) / 12
    
    pv_defender = after_tax_opportunity_cost
    flujos_defender = []
    
    for year in range(1, int(life_years_defender) + 1):
        depreciation = (book_value - float(defender.salvage_value or 0)) / life_years_defender
        tax_shield = depreciation * tax_rate
        after_tax_cash_flow = annual_op_ex_defender * (1 - tax_rate) - tax_shield
        pv = after_tax_cash_flow / ((1 + wacc) ** year)
        
        pv_defender += pv
        flujos_defender.append({
            'year': year,
            'cash_flow': annual_op_ex_defender,
            'depreciation': depreciation,
            'tax_shield': tax_shield,
            'after_tax_cash_flow': after_tax_cash_flow,
            'pv': pv
        })
    
    # Restar valor de salvamento
    pv_defender -= float(defender.salvage_value or 0) / ((1 + wacc) ** life_years_defender)
    
    eac_defender = calcular_eac(pv_defender, life_years_defender, wacc)
    
    # Análisis Challenger
    initial_cost = float(challenger.purchase_price or 0) + \
                  float(challenger.installation_and_training_cost or 0) + \
                  float(challenger.setup_costs or 0)
    
    annual_op_ex_challenger = float(challenger.annual_maintenance_labor_parts or 0) + \
                            (float(challenger.operator_labor_cost or 0) * float(challenger.monthly_operating_hours or 0) * 12)
    
    life_years_challenger = (challenger.useful_life or 180) / 12
    depreciable_base = float(challenger.purchase_price or 0) + float(challenger.installation_and_training_cost or 0)
    
    pv_challenger = initial_cost
    flujos_challenger = []
    
    for year in range(1, int(life_years_challenger) + 1):
        depreciation = (depreciable_base - float(challenger.salvage_value or 0)) / life_years_challenger
        tax_shield = depreciation * tax_rate
        after_tax_cash_flow = annual_op_ex_challenger * (1 - tax_rate) - tax_shield
        pv = after_tax_cash_flow / ((1 + wacc) ** year)
        
        pv_challenger += pv
        flujos_challenger.append({
            'year': year,
            'cash_flow': annual_op_ex_challenger,
            'depreciation': depreciation,
            'tax_shield': tax_shield,
            'after_tax_cash_flow': after_tax_cash_flow,
            'pv': pv
        })
    
    # Restar valor de salvamento
    pv_challenger -= float(challenger.salvage_value or 0) / ((1 + wacc) ** life_years_challenger)
    
    eac_challenger = calcular_eac(pv_challenger, life_years_challenger, wacc)
    
    return {
        'pv_defender': Decimal(str(pv_defender)),
        'eac_defender': Decimal(str(eac_defender)),
        'pv_challenger': Decimal(str(pv_challenger)),
        'eac_challenger': Decimal(str(eac_challenger)),
        'flujos_defender': flujos_defender,
        'flujos_challenger': flujos_challenger
    }

def calcular_eac(pv_costs, life_in_years, wacc):
    """Calcular Equivalent Annual Cost"""
    if life_in_years <= 0 or wacc <= 0:
        return pv_costs / (life_in_years or 1)
    
    annuity_factor = (1 - (1 + wacc) ** (-life_in_years)) / wacc
    if annuity_factor == 0:
        return pv_costs
    
    eac = pv_costs / annuity_factor
    return eac if abs(eac) != float('inf') else 0

def api_analisis_guardados(request):
    """API para listar análisis guardados"""
    analisis = AnalisisComparativo.objects.select_related('defender', 'challenger').all().order_by('-fecha_creacion')
    
    data = []
    for a in analisis:
        data.append({
            'id': str(a.id),
            'nombre_analisis': a.nombre_analisis,
            'defender_nombre': a.defender.nombre,
            'challenger_nombre': a.challenger.nombre,
            'eac_defender': float(a.eac_defender) if a.eac_defender else 0,
            'eac_challenger': float(a.eac_challenger) if a.eac_challenger else 0,
            'recomendacion': a.recomendacion,
            'fecha_creacion': a.fecha_creacion.isoformat(),
        })
    
    return JsonResponse(data, safe=False)

def api_analisis_detalle(request, analisis_id):
    """API para obtener detalles de un análisis específico"""
    analisis = get_object_or_404(AnalisisComparativo, id=analisis_id)
    
    flujos_caja = list(analisis.flujos_caja.all().values())
    tabla_amortizacion = list(analisis.tabla_amortizacion.all().values())
    
    data = {
        'analisis': {
            'id': str(analisis.id),
            'nombre_analisis': analisis.nombre_analisis,
            'wacc': float(analisis.wacc),
            'tax_rate': float(analisis.tax_rate),
            'financing_rate': float(analisis.financing_rate),
            'financing_months': analisis.financing_months,
            'pv_defender': float(analisis.pv_defender) if analisis.pv_defender else 0,
            'eac_defender': float(analisis.eac_defender) if analisis.eac_defender else 0,
            'pv_challenger': float(analisis.pv_challenger) if analisis.pv_challenger else 0,
            'eac_challenger': float(analisis.eac_challenger) if analisis.eac_challenger else 0,
            'recomendacion': analisis.recomendacion,
        },
        'defender': {
            'id': str(analisis.defender.id),
            'nombre': analisis.defender.nombre,
        },
        'challenger': {
            'id': str(analisis.challenger.id),
            'nombre': analisis.challenger.nombre,
        },
        'flujos_caja': flujos_caja,
        'tabla_amortizacion': tabla_amortizacion,
    }
    
    return JsonResponse(data)

def dashboard_amortizacion(request):
    """Vista para mostrar el dashboard de amortización"""
    return render(request, 'amortizacion_dashboard.html')

def analisis_lista(request):
    """Obtener lista de análisis únicos"""
    try:
        # Obtener todos los análisis únicos con información básica
        analisis = AnalisisComparativo.objects.all()
        
        data = []
        for a in analisis:
            # Calcular pago mensual desde el primer registro de la tabla
            primer_pago = TablaAmortizacion.objects.filter(analisis=a, mes=1).first()
            pago_mensual = float(primer_pago.pago_mensual) if primer_pago else 0
            
            # Calcular monto del préstamo (balance inicial del primer mes)
            monto_prestamo = float(primer_pago.balance_inicial) if primer_pago else 0
            
            data.append({
                'id': str(a.id),  # Convertir UUID a string
                'nombre_analisis': a.nombre_analisis,
                'monto_prestamo': monto_prestamo,
                'tasa_interes': float(a.financing_rate),  # ✅ Del modelo, no quemado
                'plazo_meses': a.financing_months,  # ✅ Del modelo, no quemado
                'pago_mensual': pago_mensual,
                'wacc': float(a.wacc) * 100,  # Convertir a porcentaje para mostrar
                'tax_rate': float(a.tax_rate) * 100,  # Convertir a porcentaje para mostrar
                'fecha_creacion': a.fecha_creacion.strftime('%Y-%m-%d %H:%M'),
                'recomendacion': a.recomendacion or 'Pendiente'
            })
        
        return JsonResponse({'analisis': data})
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def analisis_detalle(request, analisis_id):
    """Obtener detalles completos de amortización"""
    try:
        # Verificar que el análisis existe
        analisis = AnalisisComparativo.objects.get(id=analisis_id)
        
        # Obtener toda la tabla de amortización ordenada por mes
        tabla = TablaAmortizacion.objects.filter(analisis=analisis).order_by('mes')
        
        tabla_data = []
        for row in tabla:
            tabla_data.append({
                'mes': row.mes,
                'balance_inicial': str(row.balance_inicial),
                'pago_mensual': str(row.pago_mensual),
                'pago_principal': str(row.pago_principal),
                'pago_interes': str(row.pago_interes),
                'balance_final': str(row.balance_final)
            })
        
        # Incluir información del análisis con parámetros reales
        return JsonResponse({
            'id': str(analisis_id),
            'nombre_analisis': analisis.nombre_analisis,
            'parametros': {
                'wacc': float(analisis.wacc) * 100,
                'tax_rate': float(analisis.tax_rate) * 100,
                'financing_rate': float(analisis.financing_rate),
                'financing_months': analisis.financing_months
            },
            'tabla_amortizacion': tabla_data,
            'recomendacion': analisis.recomendacion,
            'fecha_creacion': analisis.fecha_creacion.strftime('%Y-%m-%d %H:%M')
        })
    
    except AnalisisComparativo.DoesNotExist:
        return JsonResponse({'error': 'Análisis no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    
def eliminar_analisis(request, analisis_id):
    """Eliminar un análisis específico y todos sus registros relacionados"""
    if request.method != 'DELETE':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        # Buscar el análisis
        analisis = AnalisisComparativo.objects.get(id=analisis_id)
        
        # Obtener nombre para el mensaje de confirmación
        nombre_analisis = analisis.nombre_analisis
        
        # Django automáticamente eliminará los registros relacionados debido a on_delete=models.CASCADE
        # Pero podemos hacerlo explícitamente para tener más control
        
        # 1. Eliminar registros de TablaAmortizacion
        tabla_count = TablaAmortizacion.objects.filter(analisis=analisis).count()
        TablaAmortizacion.objects.filter(analisis=analisis).delete()
        
        # 2. Eliminar registros de FlujoCaja
        flujo_count = FlujoCaja.objects.filter(analisis=analisis).count()
        FlujoCaja.objects.filter(analisis=analisis).delete()
        
        # 3. Finalmente eliminar el AnalisisComparativo
        analisis.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Análisis "{nombre_analisis}" eliminado correctamente',
            'detalles': {
                'registros_tabla_amortizacion': tabla_count,
                'registros_flujo_caja': flujo_count
            }
        })
    
    except AnalisisComparativo.DoesNotExist:
        return JsonResponse({'error': 'Análisis no encontrado'}, status=404)
    
    except Exception as e:
        return JsonResponse({
            'error': f'Error al eliminar el análisis: {str(e)}'
        }, status=500)

def confirmar_eliminacion(request, analisis_id):
    """Obtener información del análisis antes de eliminarlo"""
    try:
        analisis = AnalisisComparativo.objects.get(id=analisis_id)
        
        # Contar registros relacionados
        tabla_count = TablaAmortizacion.objects.filter(analisis=analisis).count()
        flujo_count = FlujoCaja.objects.filter(analisis=analisis).count()
        
        return JsonResponse({
            'analisis': {
                'id': str(analisis.id),
                'nombre': analisis.nombre_analisis,
                'fecha_creacion': analisis.fecha_creacion.strftime('%Y-%m-%d %H:%M'),
            },
            'registros_relacionados': {
                'tabla_amortizacion': tabla_count,
                'flujos_caja': flujo_count,
                'total': tabla_count + flujo_count
            }
        })
    
    except AnalisisComparativo.DoesNotExist:
        return JsonResponse({'error': 'Análisis no encontrado'}, status=404)
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# ==================== VISTAS DE Clientes ====================

def clientes_list(request):
    """Lista todos los clientes activos con búsqueda opcional"""
    clientes = Cliente.objects.all()
    
    # Implementar búsqueda si hay parámetro 'q'
    query = request.GET.get('q')
    if query:
        clientes = clientes.filter(
            Q(id__icontains=query) |
            Q(nombre__icontains=query) |
            Q(contacto__icontains=query)
        )
    
    return render(request, 'listar_clientes.html', {'clientes': clientes})

def clientes_crear(request):
    """Crear nuevo cliente"""
    if request.method == 'POST':
        try:
            Cliente.objects.create(
                id=request.POST.get('id'),
                nombre=request.POST.get('nombre'),
                city=request.POST.get('city'),  # Cambio: contacto -> city
                email=request.POST.get('email') or None,
                telefono=request.POST.get('telefono') or None,
                terminos_contractuales=int(request.POST.get('terminos_contractuales')),
                average_days_to_pay=int(request.POST.get('average_days_to_pay')),
                observaciones=request.POST.get('observaciones', '')
            )
            messages.success(request, 'Cliente creado exitosamente')
            return redirect('clientes_list')
        except Exception as e:
            messages.error(request, f'Error al crear cliente: {str(e)}')
    
    return render(request, 'crear_clientes.html')

def clientes_editar(request, cliente_id):  # Cambiar de 'pk' a 'cliente_id'
    """Editar cliente existente"""
    cliente = get_object_or_404(Cliente, pk=cliente_id)  # Cambiar de pk=id a pk=cliente_id
    
    if request.method == 'POST':
        try:
            cliente.nombre = request.POST.get('nombre')
            cliente.contacto = request.POST.get('contacto')
            cliente.email = request.POST.get('email') or None
            cliente.city = request.POST.get('city') or None
            cliente.telefono = request.POST.get('telefono') or None
            cliente.terminos_contractuales = int(request.POST.get('terminos_contractuales'))
            cliente.observaciones = request.POST.get('observaciones', '')
            cliente.save()
            messages.success(request, 'Cliente actualizado exitosamente')
            return redirect('clientes_list')
        except Exception as e:
            messages.error(request, f'Error al actualizar cliente: {str(e)}')
    
    return render(request, 'editar_clientes.html', {'cliente': cliente})

def clientes_eliminar(request, cliente_id):
    """Eliminar cliente existente"""
    print(f"=== DEBUG ELIMINAR CLIENTE ===")
    print(f"Método HTTP: {request.method}")  
    print(f"Cliente ID recibido: {cliente_id}")
    print(f"URL completa: {request.get_full_path()}")
    
    if request.method == 'POST':
        print("Es un POST request - procediendo con eliminación")
        try:
            cliente = get_object_or_404(Cliente, id=cliente_id)  # Sin int() si es string
            print(f"Cliente encontrado: {cliente.nombre} (ID: {cliente.id})")
            
            nombre_cliente = cliente.nombre
            cliente.delete()
            print(f"Cliente {nombre_cliente} eliminado exitosamente")
            
            messages.success(request, f'Cliente "{nombre_cliente}" eliminado exitosamente')
            return redirect('clientes_list')
            
        except Exception as e:
            print(f"Error al eliminar cliente: {str(e)}")
            messages.error(request, f'Error al eliminar cliente: {str(e)}')
            return redirect('clientes_list')
    else:
        print(f"Método no permitido: {request.method}")
        return HttpResponseNotAllowed(['POST'])

# ==================== VISTAS DE PROVEEDORES ====================

def proveedores_list(request):
    """Lista todos los proveedores activos"""
    proveedores = Proveedor.objects.filter()
    return render(request, 'listar_proveedores.html', {'proveedores': proveedores})

def proveedores_crear(request):
    """Crear nuevo proveedor"""
    if request.method == 'POST':
        try:
            Proveedor.objects.create(
                id=request.POST.get('id'),
                nombre=request.POST.get('nombre'),
                contacto=request.POST.get('contacto'),
                email=request.POST.get('email') or None,
                telefono=request.POST.get('telefono') or None,
                terminos_pago=int(request.POST.get('terminos_pago')),
                # Removido: tipo_materia_prima (no existe en el modelo)
                observaciones=request.POST.get('observaciones', '')
            )
            messages.success(request, 'Proveedor creado exitosamente')
            return redirect('proveedores_list')
        except Exception as e:
            messages.error(request, f'Error al crear proveedor: {str(e)}')
    
    return render(request, 'crear_proveedores.html')

def proveedores_editar(request, pk):
    """Editar proveedor existente"""
    proveedor = get_object_or_404(Proveedor, pk=pk)
    
    if request.method == 'POST':
        try:
            proveedor.nombre = request.POST.get('nombre')
            proveedor.contacto = request.POST.get('contacto')
            proveedor.email = request.POST.get('email') or None
            proveedor.telefono = request.POST.get('telefono') or None
            proveedor.terminos_pago = int(request.POST.get('terminos_pago'))
            proveedor.observaciones = request.POST.get('observaciones', '')
            proveedor.save()
            messages.success(request, 'Proveedor actualizado exitosamente')
            return redirect('proveedores_list')
        except Exception as e:
            messages.error(request, f'Error al actualizar proveedor: {str(e)}')
    
    return render(request, 'editar_proveedores.html', {'proveedor': proveedor})

def proveedores_eliminar(request, pk):
    """Eliminar proveedor existente"""
    print(f"=== DEBUG ELIMINAR PROVEEDOR ===")
    print(f"Método HTTP: {request.method}")  
    print(f"Proveedor PK recibido: {pk}")
    print(f"URL completa: {request.get_full_path()}")
    
    if request.method == 'POST':
        print("Es un POST request - procediendo con eliminación")
        try:
            proveedor = get_object_or_404(Proveedor, pk=pk)
            print(f"Proveedor encontrado: {proveedor.nombre} (ID: {proveedor.id})")
            
            nombre_proveedor = proveedor.nombre
            proveedor.delete()
            print(f"Proveedor {nombre_proveedor} eliminado exitosamente")
            
            messages.success(request, f'Proveedor "{nombre_proveedor}" eliminado exitosamente')
            return redirect('proveedores_list')
            
        except Exception as e:
            print(f"Error al eliminar proveedor: {str(e)}")
            messages.error(request, f'Error al eliminar proveedor: {str(e)}')
            return redirect('proveedores_list')
    else:
        print(f"Método no permitido: {request.method}")
        return HttpResponseNotAllowed(['POST'])

# ==================== VISTAS DE REGISTROS ====================

def registros_list(request):
    """
    Vista minimalista para listar registros. Los cálculos se delegan a JavaScript.
    """
    query = request.GET.get('q', '')
    
    # Query base optimizada que se ordena por defecto según el Meta del modelo
    registros_qs = Registro.objects.select_related('cliente')

    # Filtra solo si hay un término de búsqueda
    if query:
        registros_qs = registros_qs.filter(
            Q(id__icontains=query) |
            Q(cliente__nombre__icontains=query) |
            Q(estado_cobro__icontains=query)
        )

    context = {
        'registros': registros_qs,
        'query': query,
    }

    return render(request, 'listar_registros.html', context)

def registros_crear(request):
    """Vista para crear registro con gestión integral y cálculo automático de fechas."""
    
    # --- Lógica para la petición GET (Mostrar el formulario) ---
    if request.method != 'POST':
        form = RegistroForm()
        
        # 1. PREPARAR DATOS PARA EL FRONTEND (CAMBIO PRINCIPAL)
        # Se obtienen los términos de todos los clientes y proveedores para inyectarlos en el script.
        # Esto es más eficiente que hacer múltiples llamadas AJAX.
        
        clientes = Cliente.objects.filter()
        clientes_data = {
            str(c.id): {
                'nombre': c.nombre,
                'terminos': c.terminos_contractuales
            } for c in clientes
        }

        proveedores = Proveedor.objects.filter()
        proveedores_data = {
            str(p.id): {
                'nombre': p.nombre,
                'terminos': p.terminos_pago
            } for p in proveedores
        }

        context = {
            'form': form,
            'fecha_hoy': date.today().isoformat(),
            'metodos_pago': Registro.METODO_PAGO_CHOICES,
            # Convertir los diccionarios a JSON para que JS los pueda leer
            'clientes_data': json.dumps(clientes_data),
            'proveedores_data': json.dumps(proveedores_data),
        }
        # Asegúrate que el nombre del template sea el correcto
        return render(request, 'crear_registros.html', context)

    # --- Lógica para la petición POST (Guardar el formulario) ---
    form = RegistroForm(request.POST)
    
    obligaciones_data = request.POST.get('obligaciones_data', '[]')
    pagos_cliente_data = request.POST.get('pagos_cliente_data', '[]')
    pagos_proveedor_data = request.POST.get('pagos_proveedor_data', '[]')
    
    try:
        obligaciones = json.loads(obligaciones_data)
        pagos_cliente = json.loads(pagos_cliente_data)
        pagos_proveedor = json.loads(pagos_proveedor_data)
    except json.JSONDecodeError:
        messages.error(request, 'Error en el formato de datos JSON.')
        # Re-renderiza el contexto como lo haría la petición GET
        # (Esto es una mejora para que el formulario no pierda los datos de JS)
        # Puedes copiar la lógica del GET aquí si quieres que el formulario se repoble
        # con los datos de JS en caso de error. Por simplicidad, se omite por ahora.
        return redirect('registros_crear') # Redirigir es más simple

    if form.is_valid():
        try:
            with transaction.atomic():
                registro = form.save(commit=False)
                
                # 2. CALCULAR FECHAS DE VENCIMIENTO EN EL BACKEND (CAMBIO PRINCIPAL)
                # El frontend envía 'fecha_recepcion'. El backend calcula la 'fecha_vencimiento' final.
                # Esto garantiza que la lógica de negocio resida en el servidor.
                
                # Cache para proveedores para no consultar la DB en un bucle
                proveedores_cache = {str(p.id): p for p in Proveedor.objects.filter(
                    id__in=[obl.get('proveedor_id') for obl in obligaciones]
                )}

                obligaciones_procesadas = []
                for i, obligacion in enumerate(obligaciones):
                    proveedor_id = obligacion.get('proveedor_id')
                    fecha_recepcion_str = obligacion.get('fecha_recepcion')

                    if not proveedor_id or not fecha_recepcion_str:
                        raise ValidationError(f'La obligación #{i+1} debe tener proveedor y fecha de recepción.')
                    
                    proveedor = proveedores_cache.get(proveedor_id)
                    if not proveedor:
                        raise ValidationError(f'Proveedor con ID {proveedor_id} no encontrado.')

                    # Calcular la fecha de vencimiento real
                    fecha_recepcion = datetime.strptime(fecha_recepcion_str, '%Y-%m-%d').date()
                    fecha_vencimiento = fecha_recepcion + timedelta(days=proveedor.terminos_pago)
                    
                    obligacion_procesada = obligacion.copy()
                    obligacion_procesada['id'] = i + 1
                    obligacion_procesada['fecha_vencimiento'] = fecha_vencimiento.isoformat() # Guardar fecha calculada
                    obligaciones_procesadas.append(obligacion_procesada)

                # El resto de tu lógica de validación y procesamiento es correcta y se mantiene
                # ... (validaciones de pagos, asignación de IDs, etc.) ...
                total_pagos_cliente = Decimal('0.00')
                for pago in pagos_cliente:
                    total_pagos_cliente += Decimal(str(pago.get('monto', 0)))
                
                if total_pagos_cliente > registro.valor_cobrar_cliente:
                    raise ValidationError(f'Los pagos del cliente (${total_pagos_cliente:,.2f}) exceden el valor a cobrar (${registro.valor_cobrar_cliente:,.2f}).')

                # Procesar IDs
                pagos_cliente_procesados = [{'id': i + 1, **pago} for i, pago in enumerate(pagos_cliente) if pago.get('monto')]
                pagos_proveedor_procesados = [{'id': i + 1, **pago} for i, pago in enumerate(pagos_proveedor) if pago.get('monto')]
                
                # Asignar datos procesados
                registro.obligaciones_data = obligaciones_procesadas
                registro.pagos_cliente_data = pagos_cliente_procesados
                registro.pagos_proveedor_data = pagos_proveedor_procesados
                
                # El método save() del modelo se encargará de la fecha_limite_cobro del cliente
                registro.save() 
                if registro.cliente:
                    registro.cliente.actualizar_dias_promedio_pago()
                messages.success(request, f'Registro {registro.id} creado exitosamente.')
                return redirect('registros_list')
                
        except ValidationError as e:
            # Captura errores de validación de forma clara
            messages.error(request, e.message)
        except Exception as e:
            messages.error(request, f'Error inesperado al guardar: {str(e)}')
    else:
        # Si el formulario principal no es válido
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(request, f'Error en el campo "{form.fields[field].label}": {error}')
    
    # Si algo falla, volver a renderizar el formulario
    # (Aquí también podrías repoblar el contexto para no perder datos)
    return redirect('registros_crear')

def registros_editar(request, id):
    """Vista para editar registro existente con gestión integral."""
    
    # Obtener el registro o devolver 404
    registro = get_object_or_404(Registro, id=id)  # Changed registro_id to id
    
    # --- Lógica para la petición GET (Mostrar el formulario con datos existentes) ---
    if request.method != 'POST':
        form = RegistroForm(instance=registro)
        
        # Preparar datos para el frontend
        clientes = Cliente.objects.filter()
        clientes_data = {
            str(c.id): {
                'nombre': c.nombre,
                'terminos': c.terminos_contractuales
            } for c in clientes
        }

        proveedores = Proveedor.objects.filter()
        proveedores_data = {
            str(p.id): {
                'nombre': p.nombre,
                'terminos': p.terminos_pago
            } for p in proveedores
        }

        # Preparar datos existentes del registro para el frontend
        registro_data = {
            'obligaciones': registro.obtener_obligaciones(),
            'pagos_cliente': registro.obtener_pagos_cliente(),
            'pagos_proveedor': registro.obtener_pagos_proveedor(),
        }

        context = {
            'form': form,
            'registro': registro,
            'fecha_hoy': date.today().isoformat(),
            'metodos_pago': Registro.METODO_PAGO_CHOICES,
            'clientes_data': json.dumps(clientes_data),
            'proveedores_data': json.dumps(proveedores_data),
            'registro_data': json.dumps(registro_data),
            'es_edicion': True,
        }
        return render(request, 'editar_registro.html', context)

    # --- Lógica para la petición POST (Actualizar el formulario) ---
    form = RegistroForm(request.POST, instance=registro)
    
    obligaciones_data = request.POST.get('obligaciones_data', '[]')
    pagos_cliente_data = request.POST.get('pagos_cliente_data', '[]')
    pagos_proveedor_data = request.POST.get('pagos_proveedor_data', '[]')
    
    try:
        obligaciones = json.loads(obligaciones_data)
        pagos_cliente = json.loads(pagos_cliente_data)
        pagos_proveedor = json.loads(pagos_proveedor_data)
    except json.JSONDecodeError:
        messages.error(request, 'Error en el formato de datos JSON.')
        return redirect('registros_editar', id=id)  # Changed registro_id to id

    if form.is_valid():
        try:
            with transaction.atomic():
                # Guardar los cambios básicos del formulario
                registro_actualizado = form.save(commit=False)
                
                # Procesar obligaciones
                proveedores_cache = {str(p.id): p for p in Proveedor.objects.filter(
                    id__in=[obl.get('proveedor_id') for obl in obligaciones if obl.get('proveedor_id')]
                )}

                obligaciones_procesadas = []
                for i, obligacion in enumerate(obligaciones):
                    proveedor_id = obligacion.get('proveedor_id')
                    fecha_recepcion_str = obligacion.get('fecha_recepcion')

                    if not proveedor_id or not fecha_recepcion_str:
                        raise ValidationError(f'La obligación #{i+1} debe tener proveedor y fecha de recepción.')
                    
                    proveedor = proveedores_cache.get(proveedor_id)
                    if not proveedor:
                        raise ValidationError(f'Proveedor con ID {proveedor_id} no encontrado.')

                    # Calcular la fecha de vencimiento
                    fecha_recepcion = datetime.strptime(fecha_recepcion_str, '%Y-%m-%d').date()
                    fecha_vencimiento = fecha_recepcion + timedelta(days=proveedor.terminos_pago)
                    
                    obligacion_procesada = obligacion.copy()
                    # Mantener ID existente o crear nuevo
                    if 'id' not in obligacion_procesada or not obligacion_procesada['id']:
                        obligacion_procesada['id'] = i + 1
                    obligacion_procesada['fecha_vencimiento'] = fecha_vencimiento.isoformat()
                    obligacion_procesada['proveedor_nombre'] = proveedor.nombre
                    obligaciones_procesadas.append(obligacion_procesada)

                # Validar pagos del cliente
                total_pagos_cliente = Decimal('0.00')
                for pago in pagos_cliente:
                    if pago.get('monto'):
                        total_pagos_cliente += Decimal(str(pago.get('monto', 0)))
                
                if total_pagos_cliente > registro_actualizado.valor_cobrar_cliente:
                    raise ValidationError(f'Los pagos del cliente (${total_pagos_cliente:,.2f}) exceden el valor a cobrar (${registro_actualizado.valor_cobrar_cliente:,.2f}).')

                # Procesar pagos manteniendo IDs existentes o creando nuevos
                pagos_cliente_procesados = []
                for i, pago in enumerate(pagos_cliente):
                    if pago.get('monto'):
                        pago_procesado = pago.copy()
                        if 'id' not in pago_procesado or not pago_procesado['id']:
                            # Generar nuevo ID basado en los existentes
                            max_id = max([p.get('id', 0) for p in pagos_cliente_procesados], default=0)
                            pago_procesado['id'] = max_id + 1
                        pagos_cliente_procesados.append(pago_procesado)
                
                pagos_proveedor_procesados = []
                for i, pago in enumerate(pagos_proveedor):
                    if pago.get('monto'):
                        pago_procesado = pago.copy()
                        if 'id' not in pago_procesado or not pago_procesado['id']:
                            max_id = max([p.get('id', 0) for p in pagos_proveedor_procesados], default=0)
                            pago_procesado['id'] = max_id + 1
                        pagos_proveedor_procesados.append(pago_procesado)
                
                # Actualizar datos del registro
                registro_actualizado.obligaciones_data = obligaciones_procesadas
                registro_actualizado.pagos_cliente_data = pagos_cliente_procesados
                registro_actualizado.pagos_proveedor_data = pagos_proveedor_procesados
                
                # Guardar registro actualizado
                registro_actualizado.save()
                
                # Actualizar estado de cobro automáticamente
                registro_actualizado.actualizar_estado_cobro()
                
                # Actualizar días promedio de pago del cliente
                if registro_actualizado.cliente:
                    registro_actualizado.cliente.actualizar_dias_promedio_pago()
                    
                messages.success(request, f'Registro {registro_actualizado.id} actualizado exitosamente.')
                return redirect('registros_list')
                
        except ValidationError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f'Error inesperado al actualizar: {str(e)}')
    else:
        # Si el formulario principal no es válido
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(request, f'Error en el campo "{form.fields[field].label}": {error}')
    
    # Si algo falla, volver a mostrar el formulario
    return redirect('registros_editar', id=id)  # Changed registro_id to id

# En views.py
def registros_eliminar(request, registro_id):
    if request.method == 'POST':
        registro = get_object_or_404(Registro, id=registro_id)
        registro.delete()
    return redirect('registros_list')

# Vista auxiliar para validar ID de registro
def validar_id_registro(request):
    """Vista AJAX para validar si un ID de registro ya existe"""
    if request.method == 'GET':
        id_registro = request.GET.get('id', '')
        existe = Registro.objects.filter(id=id_registro).exists()
        return JsonResponse({'existe': existe})

# Vista auxiliar para obtener términos de cliente
def obtener_terminos_cliente(request, cliente_id):
    """Vista AJAX para obtener términos de un cliente"""
    try:
        cliente = get_object_or_404(Cliente, id=cliente_id)
        return JsonResponse({
            'success': True,
            'terminos_contractuales': cliente.terminos_contractuales,
            'nombre': cliente.nombre
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': 'Error al obtener datos del cliente'
        })

# Vista auxiliar para obtener términos de proveedor
def obtener_terminos_proveedor(request, proveedor_id):
    """Vista AJAX para obtener términos de un proveedor"""
    try:
        proveedor = get_object_or_404(Proveedor, id=proveedor_id)
        return JsonResponse({
            'success': True,
            'terminos_pago': getattr(proveedor, 'terminos_pago', 30),
            'nombre': proveedor.nombre
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': 'Error al obtener datos del proveedor'
        })

# ==================== FLUJO DE CAJA ====================

def flujo_caja_view(request, registro_id):
    """
    Vista principal para mostrar el flujo de caja detallado de un registro
    """
    registro = get_object_or_404(Registro, pk=registro_id)

    # Procesar obligaciones con información detallada
    obligaciones = []
    for obl in registro.obtener_obligaciones():
        pagos = registro.obtener_pagos_de_obligacion(obl.get('id'))
        pagado = sum([float(p.get('monto', 0)) for p in pagos])
        total = float(obl.get('valor_pagar', 0))
        saldo = total - pagado
        
        # Calcular días de vencimiento
        fecha_vencimiento = obl.get('fecha_vencimiento')
        dias_vencimiento = None
        if fecha_vencimiento:
            if isinstance(fecha_vencimiento, str):
                try:
                    fecha_vencimiento = datetime.strptime(fecha_vencimiento, '%Y-%m-%d').date()
                except ValueError:
                    fecha_vencimiento = None
            
            if fecha_vencimiento:
                dias_vencimiento = (fecha_vencimiento - date.today()).days

        obligaciones.append({
            'id': obl.get('id'),
            'proveedor': obl.get('proveedor_nombre', 'N/A'),
            'valor_pagar': total,
            'pagado': pagado,
            'saldo': saldo,
            'fecha_vencimiento': fecha_vencimiento,
            'dias_vencimiento': dias_vencimiento,
            'descripcion': obl.get('descripcion', ''),
        })

    # Ordenar obligaciones por fecha de vencimiento
    obligaciones.sort(key=lambda x: x['fecha_vencimiento'] if x['fecha_vencimiento'] else date.max)

    context = {
        'registro': registro,
        'obligaciones': obligaciones,
    }
    return render(request, 'flujo_caja.html', context)

def calcular_flujo_caja(request):
    """
    Vista para calcular proyecciones de flujo de caja
    """
    registro_id = request.GET.get('registro_id')
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    registro = get_object_or_404(Registro, id=registro_id)

    # Convertir fechas
    try:
        fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
    except (TypeError, ValueError):
        fecha_inicio = registro.fecha_entrega_cliente

    try:
        fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
    except (TypeError, ValueError):
        fecha_fin = date.today() + timezone.timedelta(days=30)  # 30 días hacia adelante por defecto

    # Proyección diaria de flujo
    proyecciones = registro.obtener_proyeccion_flujo(fecha_inicio, fecha_fin)

    # Agrupar ingresos y egresos por fecha
    flujo_por_fecha = {}
    for proyeccion in proyecciones:
        fecha = proyeccion['fecha']
        monto = Decimal(proyeccion['monto'])
        tipo = proyeccion['tipo']
        concepto = proyeccion['concepto']

        if fecha not in flujo_por_fecha:
            flujo_por_fecha[fecha] = {
                'fecha': fecha.isoformat(),
                'ingresos_esperados': 0,
                'egresos_esperados': 0,
                'detalles': {
                    'ingresos': [],
                    'egresos': []
                }
            }

        if tipo == 'ingreso':
            flujo_por_fecha[fecha]['ingresos_esperados'] += float(monto)
            flujo_por_fecha[fecha]['detalles']['ingresos'].append({
                'cliente': registro.cliente.nombre,
                'concepto': concepto,
                'monto': float(monto)
            })
        else:
            flujo_por_fecha[fecha]['egresos_esperados'] += float(monto)
            flujo_por_fecha[fecha]['detalles']['egresos'].append({
                'proveedor': concepto.replace("Pago a ", ""),
                'concepto': concepto,
                'monto': float(monto)
            })

    # Ordenar por fecha y calcular flujo acumulado
    fechas_ordenadas = sorted(flujo_por_fecha.keys())
    flujo_ordenado = []
    flujo_acumulado = 0
    dias_con_flujo_negativo = 0

    for fecha in fechas_ordenadas:
        dia = flujo_por_fecha[fecha]
        flujo_neto = dia['ingresos_esperados'] - dia['egresos_esperados']
        flujo_acumulado += flujo_neto
        
        if flujo_acumulado < 0:
            dias_con_flujo_negativo += 1

        dia['flujo_neto'] = flujo_neto
        dia['flujo_acumulado'] = flujo_acumulado
        flujo_ordenado.append(dia)

    # Calcular resumen estadístico
    resumen = {
        'total_ingresos': sum(d['ingresos_esperados'] for d in flujo_ordenado),
        'total_egresos': sum(d['egresos_esperados'] for d in flujo_ordenado),
        'flujo_neto_total': sum(d['flujo_neto'] for d in flujo_ordenado),
        'flujo_acumulado_min': min((d['flujo_acumulado'] for d in flujo_ordenado), default=0),
        'flujo_acumulado_max': max((d['flujo_acumulado'] for d in flujo_ordenado), default=0),
        'dias_con_flujo_negativo': dias_con_flujo_negativo,
        'periodo_dias': (fecha_fin - fecha_inicio).days + 1,
        'promedio_diario': sum(d['flujo_neto'] for d in flujo_ordenado) / len(flujo_ordenado) if flujo_ordenado else 0
    }

    return JsonResponse({
        'success': True,
        'data': flujo_ordenado,
        'resumen': resumen,
        'registro': {
            'id': registro.id,
            'cliente': registro.cliente.nombre,
            'valor_cobrar': float(registro.valor_cobrar_cliente),
            'saldo_pendiente': float(registro.calcular_saldo_pendiente_cliente()),
        }
    })

def obtener_datos_dashboard(request, registro_id):
    """
    Vista para obtener datos del dashboard en tiempo real (AJAX)
    """
    registro = get_object_or_404(Registro, pk=registro_id)
    
    # Calcular métricas principales
    valor_cobrar = float(registro.valor_cobrar_cliente)
    saldo_pendiente = float(registro.calcular_saldo_pendiente_cliente())
    total_obligaciones = float(registro.calcular_total_obligaciones())
    margen_bruto = float(registro.margen_bruto) if hasattr(registro, 'margen_bruto') else 0
    
    # Calcular porcentajes
    porcentaje_cobrado = ((valor_cobrar - saldo_pendiente) / valor_cobrar * 100) if valor_cobrar > 0 else 0
    
    # Calcular porcentaje pagado a proveedores
    pagado_proveedores = 0
    for obl in registro.obtener_obligaciones():
        pagos = registro.obtener_pagos_de_obligacion(obl.get('id'))
        pagado_proveedores += sum([float(p.get('monto', 0)) for p in pagos])
    
    porcentaje_pagado_proveedores = (pagado_proveedores / total_obligaciones * 100) if total_obligaciones > 0 else 0
    
    # Análisis de riesgo
    riesgo = registro.analizar_riesgo_cobro() if hasattr(registro, 'analizar_riesgo_cobro') else {
        'nivel': 'bajo',
        'mensaje': 'Análisis de riesgo no disponible'
    }
    
    return JsonResponse({
        'success': True,
        'metricas': {
            'valor_cobrar': valor_cobrar,
            'saldo_pendiente': saldo_pendiente,
            'total_obligaciones': total_obligaciones,
            'margen_bruto': margen_bruto,
            'porcentaje_cobrado': round(porcentaje_cobrado, 1),
            'porcentaje_pagado_proveedores': round(porcentaje_pagado_proveedores, 1),
        },
        'riesgo': riesgo,
        'timeline': {
            'pagos_cliente': len(registro.obtener_pagos_cliente()),
            'pagos_proveedor': len(registro.obtener_pagos_proveedor()),
            'obligaciones_total': len(registro.obtener_obligaciones()),
        }
    })

def exportar_reporte_flujo(request, registro_id):
    """
    Vista para exportar el reporte de flujo de caja
    """
    registro = get_object_or_404(Registro, pk=registro_id)
    
    # Aquí implementarías la lógica para generar PDF, Excel, etc.
    # Por ahora retornamos un JSON con los datos
    
    obligaciones = []
    for obl in registro.obtener_obligaciones():
        pagos = registro.obtener_pagos_de_obligacion(obl.get('id'))
        pagado = sum([float(p.get('monto', 0)) for p in pagos])
        total = float(obl.get('valor_pagar', 0))
        
        obligaciones.append({
            'proveedor': obl.get('proveedor_nombre', 'N/A'),
            'valor_total': total,
            'pagado': pagado,
            'saldo': total - pagado,
            'fecha_vencimiento': str(obl.get('fecha_vencimiento', '')),
            'descripcion': obl.get('descripcion', ''),
        })
    
    data = {
        'registro': {
            'id': registro.id,
            'cliente': registro.cliente.nombre,
            'fecha_entrega': str(registro.fecha_entrega_cliente),
            'valor_cobrar': float(registro.valor_cobrar_cliente),
            'saldo_pendiente': float(registro.calcular_saldo_pendiente_cliente()),
        },
        'obligaciones': obligaciones,
        'pagos_cliente': [
            {
                'fecha': str(p.get('fecha_pago', '')),
                'monto': float(p.get('monto', 0)),
                'metodo': p.get('metodo_pago', ''),
                'referencia': p.get('referencia', ''),
            }
            for p in registro.obtener_pagos_cliente()
        ],
        'pagos_proveedor': [
            {
                'fecha': str(p.get('fecha_pago', '')),
                'monto': float(p.get('monto', 0)),
                'metodo': p.get('metodo_pago', ''),
                'obligacion_id': p.get('obligacion_id', ''),
            }
            for p in registro.obtener_pagos_proveedor()
        ],
    }
    
    return JsonResponse({
        'success': True,
        'data': data,
        'message': 'Reporte generado exitosamente'
    })

# ================= IMPORTAR REGISTROS ==================

def cargar_excel_completo(request):
    if request.method == "POST" and request.FILES.get("archivo_excel"):
        archivo = request.FILES["archivo_excel"]
        try:
            wb = openpyxl.load_workbook(archivo)

            # === 1. CLIENTES ===
            if "Clientes" in wb.sheetnames:
                hoja_clientes = wb["Clientes"]
                for fila in list(hoja_clientes.iter_rows(min_row=2, values_only=True)):
                    (id_cliente, nombre, city, email, telefono,
                     terminos, avg_dias, activo, observaciones) = fila
                    if not Cliente.objects.filter(id=id_cliente).exists():
                        Cliente.objects.create(
                            id=id_cliente,
                            nombre=nombre,
                            city=city,
                            email=email,
                            telefono=telefono,
                            terminos_contractuales=int(terminos),
                            average_days_to_pay=int(avg_dias),
                            activo=bool(activo),
                            observaciones=observaciones or ""
                        )

            # === 2. PROVEEDORES ===
            if "Proveedores" in wb.sheetnames:
                hoja_prov = wb["Proveedores"]
                for fila in list(hoja_prov.iter_rows(min_row=2, values_only=True)):
                    (id_prov, nombre, contacto, email, telefono,
                     terminos, tipo_mp, activo, observaciones) = fila
                    if not Proveedor.objects.filter(id=id_prov).exists():
                        Proveedor.objects.create(
                            id=id_prov,
                            nombre=nombre,
                            contacto=contacto,
                            email=email,
                            telefono=telefono,
                            terminos_pago=int(terminos),
                            tipo_materia_prima=tipo_mp,
                            activo=bool(activo),
                            observaciones=observaciones or ""
                        )

            # === 3. REGISTROS ===
            hoja_registros = wb["Registros"]
            for fila in list(hoja_registros.iter_rows(min_row=2, values_only=True)):
                id_registro, cliente_id, fecha_entrega, valor_cobrar, estado_cobro, observaciones = fila
                try:
                    cliente = Cliente.objects.get(id=cliente_id)
                    if not Registro.objects.filter(id=id_registro).exists():
                        Registro.objects.create(
                            id=id_registro,
                            cliente=cliente,
                            fecha_entrega_cliente=fecha_entrega if isinstance(fecha_entrega, datetime) else datetime.strptime(fecha_entrega, "%Y-%m-%d").date(),
                            valor_cobrar_cliente=Decimal(valor_cobrar),
                            estado_cobro=estado_cobro,
                            observaciones=observaciones or ""
                        )
                except Exception as e:
                    messages.error(request, f"Registro {id_registro}: {str(e)}")

            # === 4. OBLIGACIONES ===
            hoja_obl = wb["Obligaciones"]
            for fila in list(hoja_obl.iter_rows(min_row=2, values_only=True)):
                registro_id, proveedor_id, proveedor_nombre, valor_pagar, fecha_vencimiento, descripcion, referencia = fila
                try:
                    registro = Registro.objects.get(id=registro_id)
                    registro.agregar_obligacion(
                        proveedor_nombre=proveedor_nombre,
                        valor_pagar=Decimal(valor_pagar),
                        fecha_vencimiento=fecha_vencimiento,
                        proveedor_id=proveedor_id,
                        descripcion=descripcion or "",
                        referencia=referencia or ""
                    )
                except Exception as e:
                    messages.error(request, f"Obligación para registro {registro_id}: {str(e)}")

            # === 5. PAGOS CLIENTE ===
            hoja_pc = wb["Pagos_Cliente"]
            for fila in list(hoja_pc.iter_rows(min_row=2, values_only=True)):
                registro_id, monto, fecha_pago, metodo_pago, referencia, observaciones = fila
                try:
                    registro = Registro.objects.get(id=registro_id)
                    registro.agregar_pago_cliente(
                        monto=Decimal(monto),
                        fecha_pago=fecha_pago,
                        metodo_pago=metodo_pago or "transferencia",
                        referencia=referencia or "",
                        observaciones=observaciones or ""
                    )
                except Exception as e:
                    messages.error(request, f"Pago cliente {registro_id}: {str(e)}")

            # === 6. PAGOS PROVEEDOR ===
            hoja_pp = wb["Pagos_Proveedor"]
            for fila in list(hoja_pp.iter_rows(min_row=2, values_only=True)):
                registro_id, obligacion_id, monto, fecha_pago, metodo_pago, referencia, observaciones = fila
                try:
                    registro = Registro.objects.get(id=registro_id)
                    registro.agregar_pago_proveedor(
                        obligacion_id=int(obligacion_id),
                        monto=Decimal(monto),
                        fecha_pago=fecha_pago,
                        metodo_pago=metodo_pago or "transferencia",
                        referencia=referencia or "",
                        observaciones=observaciones or ""
                    )
                except Exception as e:
                    messages.error(request, f"Pago proveedor {registro_id}: {str(e)}")

            messages.success(request, "Archivo importado correctamente.")

        except Exception as e:
            messages.error(request, f"Error general: {str(e)}")

        return redirect("cargar_excel_completo")

    return render(request, "cargar_excel_completo.html")

# ================= CXC ==================

def cuentas_por_cobrar(request):
    """Vista para Cuentas por Cobrar (CxC)"""
    
    # Obtener todos los registros con sus clientes
    registros = Registro.objects.select_related('cliente').all()
    
    cxc_data = []
    
    # Totales para el resumen
    total_original_amount = Decimal('0')
    total_net_balance = Decimal('0')
    total_not_due = Decimal('0')
    total_0_30 = Decimal('0')
    total_31_60 = Decimal('0')
    total_61_90 = Decimal('0')
    total_91_120 = Decimal('0')
    total_120_plus = Decimal('0')
    total_pagado = Decimal('0')
    
    for registro in registros:
        # Calcular saldo pendiente
        saldo_pendiente = registro.calcular_saldo_pendiente_cliente()
        
        # Calcular total pagado por el cliente
        pagos_realizados = sum(
            Decimal(str(pago.get('monto', 0))) 
            for pago in registro.obtener_pagos_cliente()
        )
        
        # Calcular días de vencimiento
        dias_vencimiento = registro.dias_vencimiento or 0
        overdue_days = abs(dias_vencimiento) if dias_vencimiento < 0 else 0
        
        # Calcular días desde la entrega
        invoice_days = (date.today() - registro.fecha_entrega_cliente).days
        
        # Distribuir el saldo en rangos de días
        not_due = Decimal('0')
        days_0_30 = Decimal('0')
        days_31_60 = Decimal('0')
        days_61_90 = Decimal('0')
        days_91_120 = Decimal('0')
        days_120_plus = Decimal('0')
        
        if saldo_pendiente > 0:
            if dias_vencimiento > 0:  # No vencido
                not_due = saldo_pendiente
            elif overdue_days <= 30:  # 0-30 días vencido
                days_0_30 = saldo_pendiente
            elif overdue_days <= 60:  # 31-60 días vencido
                days_31_60 = saldo_pendiente
            elif overdue_days <= 90:  # 61-90 días vencido
                days_61_90 = saldo_pendiente
            elif overdue_days <= 120:  # 91-120 días vencido
                days_91_120 = saldo_pendiente
            else:  # +120 días vencido
                days_120_plus = saldo_pendiente
        
        # Determinar términos de pago del cliente
        payment_terms = f"{registro.cliente.terminos_contractuales} días" if registro.cliente else "N/A"
        
        cxc_item = {
            'customer': registro.cliente.id if registro.cliente else '',
            'customer_name': registro.cliente.nombre if registro.cliente else 'Sin Cliente',
            'company_code': '1000',  # Valor fijo según necesidades
            'document_type': 'ZF',   # Valor fijo según necesidades
            'document_number': registro.id,
            'posting_date': registro.fecha_entrega_cliente,
            'document_date': registro.fecha_entrega_cliente,
            'net_due_date': registro.fecha_limite_cobro,
            'invoice_days': invoice_days,
            'overdue_days': overdue_days,
            'original_amount': registro.valor_cobrar_cliente,
            'net_balance': saldo_pendiente,
            'payment_terms': payment_terms,
            'reference': registro.id,
            'customer_reference': '',
            'not_due': not_due,
            'days_0_30': days_0_30,
            'days_31_60': days_31_60,
            'days_61_90': days_61_90,
            'days_91_120': days_91_120,
            'days_120_plus': days_120_plus,
            'estado_cobro': registro.estado_cobro,
            'esta_vencido': registro.esta_vencido,
            'pagos_realizados': pagos_realizados,
        }
        
        cxc_data.append(cxc_item)
        
        # Acumular totales
        total_original_amount += registro.valor_cobrar_cliente
        total_net_balance += saldo_pendiente
        total_not_due += not_due
        total_0_30 += days_0_30
        total_31_60 += days_31_60
        total_61_90 += days_61_90
        total_91_120 += days_91_120
        total_120_plus += days_120_plus
        total_pagado += pagos_realizados
    
    # Preparar resumen
    resumen_cxc = {
        'total_original_amount': total_original_amount,
        'total_net_balance': total_net_balance,
        'total_pagado': total_pagado,
        'total_not_due': total_not_due,
        'total_0_30': total_0_30,
        'total_31_60': total_31_60,
        'total_61_90': total_61_90,
        'total_91_120': total_91_120,
        'total_120_plus': total_120_plus,
        'porcentaje_cobrado': (total_pagado / total_original_amount * 100) if total_original_amount > 0 else 0,
        'total_registros': len(cxc_data),
    }
    
    context = {
        'cxc_data': cxc_data,
        'resumen': resumen_cxc,
        'fecha_reporte': timezone.now(),
    }
    
    return render(request, 'cuentas_por_cobrar.html', context)

def cuentas_por_pagar(request):
    """Vista para renderizar la página de Cuentas por Pagar."""
    return render(request, 'cuentas_por_pagar.html') # Asegúrate de que la ruta sea correcta

def clasificar_pagos_por_antiguedad(pagos_obligacion, fecha_inicio_obligacion):
    """
    Clasifica los pagos de una obligación en rangos de tiempo basados en
    cuándo se realizaron después de la fecha de inicio.
    """
    rangos = {
        'pagos_0_30': 0.0,
        'pagos_31_60': 0.0,
        'pagos_61_90': 0.0,
        'pagos_91_120': 0.0,
        'pagos_120_plus': 0.0,
    }

    if not fecha_inicio_obligacion:
        return rangos # No se puede calcular sin fecha de inicio

    for pago in pagos_obligacion:
        try:
            monto_pago = Decimal(str(pago.get('monto', 0)))
            fecha_pago_str = pago.get('fecha_pago')
            if not fecha_pago_str:
                continue

            fecha_pago = datetime.strptime(fecha_pago_str, '%Y-%m-%d').date()
            dias_transcurridos = (fecha_pago - fecha_inicio_obligacion).days

            if 0 <= dias_transcurridos <= 30:
                rangos['pagos_0_30'] += float(monto_pago)
            elif 31 <= dias_transcurridos <= 60:
                rangos['pagos_31_60'] += float(monto_pago)
            elif 61 <= dias_transcurridos <= 90:
                rangos['pagos_61_90'] += float(monto_pago)
            elif 91 <= dias_transcurridos <= 120:
                rangos['pagos_91_120'] += float(monto_pago)
            elif dias_transcurridos > 120:
                rangos['pagos_120_plus'] += float(monto_pago)
        except (ValueError, TypeError):
            continue
            
    return rangos

# --- API Endpoint ACTUALIZADA ---
def cuentas_por_pagar_api(request):
    """API endpoint para obtener los datos consolidados de Cuentas por Pagar."""
    try:
        registros = Registro.objects.all()
        cxp_data = []

        for registro in registros:
            for obligacion in registro.obtener_obligaciones():
                valor_original = Decimal(str(obligacion.get('valor_pagar', 0)))
                pagos_obligacion = registro.obtener_pagos_de_obligacion(obligacion.get('id'))
                pagos_realizados = sum(Decimal(str(pago.get('monto', 0))) for pago in pagos_obligacion)
                saldo_pendiente = valor_original - pagos_realizados
                
                if saldo_pendiente <= 0:
                    continue
                
                # --- Lógica de Vencimiento (para la columna 'Estado') ---
                fecha_vencimiento_str = obligacion.get('fecha_vencimiento')
                overdue_days = 0
                net_due_date = None
                if fecha_vencimiento_str:
                    try:
                        fecha_vencimiento = datetime.strptime(fecha_vencimiento_str, '%Y-%m-%d').date()
                        net_due_date = fecha_vencimiento.isoformat()
                        dias_vencidos = (date.today() - fecha_vencimiento).days
                        overdue_days = dias_vencidos if dias_vencidos > 0 else 0
                    except (ValueError, TypeError): pass

                # --- Lógica de Pagos por Antigüedad (la nueva lógica) ---
                fecha_inicio_str = obligacion.get('fecha_creacion', registro.fecha_creacion.date().isoformat())
                fecha_inicio = None
                try:
                    fecha_inicio = datetime.fromisoformat(fecha_inicio_str).date()
                except: pass

                rangos_de_pagos = clasificar_pagos_por_antiguedad(pagos_obligacion, fecha_inicio)

                # Construir el objeto para la API
                cxp_item = {
                    'id': f"{registro.id}-{obligacion.get('id')}",
                    'vendorName': obligacion.get('proveedor_nombre', 'Desconocido'),
                    'documentNumber': f"FAC-{registro.id}-{obligacion.get('id')}",
                    'postingDate': fecha_inicio.isoformat() if fecha_inicio else None,
                    'netDueDate': net_due_date,
                    'originalAmount': float(valor_original),
                    'netBalance': float(saldo_pendiente),
                    'paidAmount': float(pagos_realizados),
                    'overdueDays': overdue_days,
                    'isOverdue': overdue_days > 0,
                    'description': obligacion.get('descripcion', 'N/A'),
                    **rangos_de_pagos  # Añadir los rangos de pagos
                }
                cxp_data.append(cxp_item)

        response_data = {'success': True, 'cxp_data': cxp_data, 'fecha_reporte': timezone.now().isoformat()}
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

def cuentas_por_cobrar(request):
    """Vista para renderizar la página de Cuentas por Cobrar."""
    # El nombre de la plantilla debe coincidir con el archivo que crees.
    return render(request, 'cuentas_por_cobrar.html')

# --- NUEVA FUNCIÓN DE AYUDA PARA CLASIFICAR COBROS ---
def clasificar_cobros_por_antiguedad(pagos_cliente, fecha_inicio):
    """
    Clasifica los cobros (pagos del cliente) en rangos de tiempo basados en
    cuándo se recibieron, contando desde la fecha de inicio (Día 0).
    """
    rangos = {
        'cobros_0_30': 0.0,
        'cobros_31_60': 0.0,
        'cobros_61_90': 0.0,
        'cobros_91_120': 0.0,
        'cobros_120_plus': 0.0,
    }

    if not fecha_inicio:
        return rangos # No se puede calcular sin la fecha de entrega

    for pago in pagos_cliente:
        try:
            monto_cobrado = Decimal(str(pago.get('monto', 0)))
            fecha_pago_str = pago.get('fecha_pago')
            if not fecha_pago_str:
                continue

            fecha_pago = datetime.strptime(fecha_pago_str, '%Y-%m-%d').date()
            dias_transcurridos = (fecha_pago - fecha_inicio).days

            if 0 <= dias_transcurridos <= 30:
                rangos['cobros_0_30'] += float(monto_cobrado)
            elif 31 <= dias_transcurridos <= 60:
                rangos['cobros_31_60'] += float(monto_cobrado)
            elif 61 <= dias_transcurridos <= 90:
                rangos['cobros_61_90'] += float(monto_cobrado)
            elif 91 <= dias_transcurridos <= 120:
                rangos['cobros_91_120'] += float(monto_cobrado)
            elif dias_transcurridos > 120:
                rangos['cobros_120_plus'] += float(monto_cobrado)
        except (ValueError, TypeError):
            continue
            
    return rangos

# --- API ENDPOINT PARA CUENTAS POR COBRAR (CORREGIDA) ---
def cuentas_por_cobrar_api(request):
    """API endpoint para obtener los datos consolidados de Cuentas por Cobrar."""
    try:
        registros = Registro.objects.select_related('cliente').all()
        cxc_data = []
        
        resumen = {
            'total_facturado': Decimal('0'),
            'total_saldo_pendiente': Decimal('0'),
            'total_cobrado': Decimal('0'),
        }

        for registro in registros:
            saldo_pendiente = registro.calcular_saldo_pendiente_cliente()

            if saldo_pendiente <= 0:
                continue

            valor_original = registro.valor_cobrar_cliente
            total_cobrado = valor_original - saldo_pendiente
            
            resumen['total_facturado'] += valor_original
            resumen['total_saldo_pendiente'] += saldo_pendiente
            resumen['total_cobrado'] += total_cobrado
            
            dias_vencidos = 0
            if registro.esta_vencido:
                # El método dias_vencimiento devuelve un número negativo para días vencidos
                dias_vencidos = abs(registro.dias_vencimiento)

            # --- LÓGICA CORREGIDA ---
            # El día 0 es la fecha de entrega al cliente
            fecha_inicio = registro.fecha_entrega_cliente
            pagos_del_cliente = registro.obtener_pagos_cliente()
            rangos_cobros = clasificar_cobros_por_antiguedad(pagos_del_cliente, fecha_inicio)
            
            cxc_item = {
                'registro_id': registro.id,
                'cliente_nombre': registro.cliente.nombre if registro.cliente else "N/A",
                'fecha_entrega': fecha_inicio.isoformat() if fecha_inicio else None,
                'fecha_vencimiento': registro.fecha_limite_cobro.isoformat() if registro.fecha_limite_cobro else None,
                'dias_vencidos': dias_vencidos,
                'esta_vencido': registro.esta_vencido,
                'valor_original': float(valor_original),
                'saldo_pendiente': float(saldo_pendiente),
                'total_cobrado': float(total_cobrado),
                'estado_cobro': registro.get_estado_cobro_display(),
                **rangos_cobros # Desempacar el diccionario de cobros clasificados
            }
            cxc_data.append(cxc_item)
            
        response_data = {
            'success': True,
            'cxc_data': cxc_data,
            'resumen': {k: float(v) for k, v in resumen.items()},
            'fecha_reporte': timezone.now().isoformat(),
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
